import openpyxl

y=2

file_location='Test.xlsx'
workbook = openpyxl.load_workbook(file_location)
worksheet = workbook.worksheets[0]
worksheet.insert_cols(9)

cell_title = worksheet.cell(row=1, column=9)

cell_title = worksheet.cell(row=1, column=9)
cell_title.value = 'New Column'

for x in range(10):
    cell_to_write = worksheet.cell(row=y, column=9)
    cell_to_write.value = x
    y += 1

workbook.save('C:/Users/hp/Desktop/LinkedIn_DataCollection/Test.xlsx')